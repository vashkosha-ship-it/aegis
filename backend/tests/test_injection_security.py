"""Тесты на инъекции: Excel-формулы и выход за пределы каталога хранилища."""
from __future__ import annotations

import io

import pytest

from tests.conftest import auth_headers, make_user


class TestExcelFormulaInjection:
    """Ячейка, начинающаяся с =, +, -, @, исполняется при открытии файла."""

    async def test_malicious_full_name_is_escaped(self, client, db, admin_user):
        """Выгрузка строится по дочитанным книгам, поэтому нужен MyListEntry."""
        from app.models.book import Book
        from app.models.library import MyListEntry, MyListStatus

        payload = '=cmd|\' /C calc\'!A0'
        evil = await make_user(db, username="evil", full_name=payload)

        book = Book(title="=HYPERLINK(\"http://evil\")", author="A", description="")
        db.add(book)
        await db.commit()
        await db.refresh(book)

        db.add(MyListEntry(
            user_id=evil.id, book_id=book.id, status=MyListStatus.COMPLETED
        ))
        await db.commit()

        r = await client.get("/admin/export/reading", headers=auth_headers(admin_user))
        assert r.status_code == 200

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(r.content))
        found = False
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str) and payload in cell:
                        found = True
                        assert cell.startswith("'"), (
                            f"Формула не обезврежена: {cell!r}"
                        )
        assert found, "Тестовое значение не попало в выгрузку — проверь фикстуру"


class TestAccelPathValidation:
    """X-Accel-Redirect: ключ из БД не должен выводить за каталог storage."""

    @pytest.mark.parametrize(
        "bad_key",
        [
            "../../../../etc/passwd",
            "/etc/passwd",
            "books/pdf/../../../../etc/shadow",
            "..",
        ],
    )
    def test_traversal_rejected(self, bad_key):
        from fastapi import HTTPException

        from app.api.books import _accel_path_for_key

        with pytest.raises(HTTPException):
            _accel_path_for_key(bad_key)

    def test_normal_key_allowed(self):
        from app.api.books import _accel_path_for_key

        path = _accel_path_for_key("books/pdf/abc123.pdf")
        assert path == "/_protected_pdf/books/pdf/abc123.pdf"

    def test_special_chars_are_escaped(self):
        from app.api.books import _accel_path_for_key

        path = _accel_path_for_key("books/pdf/файл с пробелом.pdf")
        assert " " not in path
        assert path.startswith("/_protected_pdf/")
