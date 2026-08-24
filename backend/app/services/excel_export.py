"""Выгрузка отчётов в Excel.

Вынесено из роутера: сборка книги Excel — самостоятельная задача в сотню
строк, не связанная с HTTP. Здесь же живёт защита от formula injection.
"""
from __future__ import annotations

import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

# Символы, с которых Excel и LibreOffice начинают трактовать содержимое ячейки
# как формулу. ФИО и названия книг вводят пользователи, поэтому значение
# нужно обезвредить, иначе выгрузку можно превратить в средство атаки на
# машину администратора.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def xl_safe(value):
    """Обезвредить значение ячейки перед записью в файл."""
    if not isinstance(value, str):
        return value
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def build_reading_report(rows: list[tuple], summary: dict[int, dict]) -> io.BytesIO:
    """Собрать книгу Excel с отчётом о прочитанном.

    rows — записи (пользователь, книга, дата завершения);
    summary — агрегат «сколько книг прочёл каждый».

    Возвращает готовый буфер, установленный на начало.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F2B7B")

    def _write_header(sheet, names: list[str]) -> None:
        for col, name in enumerate(names, start=1):
            cell = sheet.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Лист 1: сводка по пользователям
    ws1 = wb.active
    ws1.title = "Сводка"
    _write_header(ws1, ["ФИО", "Подразделение", "Прочитано книг"])
    for i, s in enumerate(sorted(summary.values(), key=lambda x: -x["count"]), start=2):
        ws1.cell(row=i, column=1, value=xl_safe(s["fio"]))
        ws1.cell(row=i, column=2, value=xl_safe(s["dept"]))
        ws1.cell(row=i, column=3, value=s["count"])
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 16

    # Лист 2: детализация (каждая прочитанная книга)
    ws2 = wb.create_sheet("Детализация")
    _write_header(ws2, ["ФИО", "Подразделение", "Книга", "Дата завершения"])
    for i, (user, book, when) in enumerate(rows, start=2):
        ws2.cell(row=i, column=1, value=xl_safe(user.full_name or user.username))
        ws2.cell(row=i, column=2, value=xl_safe(user.department or "—"))
        ws2.cell(row=i, column=3, value=xl_safe(book.title))
        ws2.cell(row=i, column=4, value=when.strftime("%Y-%m-%d %H:%M") if when else "")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 45
    ws2.column_dimensions["D"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def report_filename(date_from: str | None, date_to: str | None) -> str:
    """Имя файла выгрузки с периодом, если он задан."""
    period = ""
    if date_from or date_to:
        period = f"_{date_from or 'нач'}_{date_to or 'кон'}"
    return f"aegis_reading{period}.xlsx"


def parse_period(
    date_from: str | None, date_to: str | None
) -> tuple[datetime | None, datetime | None]:
    """Разобрать границы периода из ISO-строк. ValueError при плохом формате."""
    from datetime import UTC

    df = dt = None
    if date_from:
        df = datetime.fromisoformat(date_from).replace(tzinfo=UTC)
    if date_to:
        dt = datetime.fromisoformat(date_to).replace(tzinfo=UTC)
    return df, dt
